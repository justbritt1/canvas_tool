" 10% tool for canvas "
" Created By: Brittany Smith '2025'"
" Sorts and cleans csv file for census report"




import os
import pandas as pd
import xlsxwriter
import csv
import openpyxl
import re
#import webview
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from io import BytesIO
from openpyxl import load_workbook
from zipfile import BadZipFile


app = Flask(__name__)

# Set a secret key for Flask
app.secret_key = 'your_secret_key_here'

#webview.create_window("File manipulation", app)

# Name of the directory for uploads
directory_name = "uploads"

# Folder path
UPLOAD_FOLDER = directory_name

# Name of the directory for downloads
DOWNLOAD_FOLDER = "census_correct"

INCORRECT_FOLDER = 'census_incorrect'

# Get the current working directory
current_directory = os.getcwd()

# Combine the current directory with the new directory names
upload_directory_path = os.path.join(current_directory, UPLOAD_FOLDER)
download_directory_path = os.path.join(current_directory, DOWNLOAD_FOLDER)
incorrect_directory_path = os.path.join(current_directory, INCORRECT_FOLDER)


# Create the directories if they don't exist
os.makedirs(upload_directory_path, exist_ok=True)
os.makedirs(download_directory_path, exist_ok=True)
os.makedirs(incorrect_directory_path, exist_ok=True)

# Set these folders in Flask's configuration (optional)
app.config['UPLOAD_FOLDER'] = upload_directory_path
app.config['DOWNLOAD_FOLDER'] = download_directory_path
app.config['INCORRECT_FOLDER'] = incorrect_directory_path


# initiate the resulting of saving file, now it's empty
save = ''
file = ''

# Define columns to keep
COLUMNS_TO_KEEP = [
    'Name', 'SISID', 'SectionSISIDs',
'Submitted','OverallScore', 'Assignment'
]

df = pd.DataFrame()  # Initialize df here

# Define the correct filename pattern/ format
FILENAME_PATTERN = re.compile(r'^\d{4}(SP|FA|SU|WI)_[A-Z]{3}_\d{3}_\d{4}_Census$', re.IGNORECASE)

# Upload page with options
@app.route('/', methods=['GET', 'POST'])
def index():
    global file_name, df

    # Define required columns
    required_columns = [
        'Name', 'SISID', 'SectionSISIDs',
'Submitted','OverallScore']

    
    valid_file_uploaded = False  # Flag to indicate if a valid file has been uploaded

    if request.method == 'POST':
        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            file_name = file.filename.rsplit('.', 1)[0]  # Extract filename without extension

            if not FILENAME_PATTERN.match(file_name):
                flash(f'Incorrect filename format: {file.filename}', 'error')
                
                # Save the incorrect file in the census_incorrect folder
                incorrect_file_path = os.path.join(app.config['INCORRECT_FOLDER'], secure_filename(file.filename))
                file.save(incorrect_file_path)

                flash(f'File saved in census_incorrect folder due to incorrect filename format.', 'success')
                return redirect(url_for('index'))  # Stop processing

            # Check if the file has an allowed extension
            allowed_extensions = {'csv', 'xlsx', 'xls', 'txt'}

            if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                try:
                    # Handle different file formats
                    if file.filename.endswith('.csv'):
                        df_temp = pd.read_csv(file)
                    elif file.filename.endswith('.xlsx'):
                        df_temp = pd.read_excel(file)
                    elif file.filename.endswith('.xls'):  # Handle .xls files
                        try:
                            # Read the .xls file using pandas
                            df_temp = pd.read_excel(file, engine='xlrd')
                            
                        except Exception as e:
                            # If Excel reading fails, it might be a text file saved as .xls
                            try:
                                file.seek(0)  # Reset the file pointer to the beginning
                                # Attempt to read the .xls as a text file
                                df_temp = pd.read_csv(file, delimiter='\t', encoding='utf-16')  # Adjust encoding if needed
                                flash('.xls file has been converted to .csv!', 'success')
                            except Exception as e:
                                flash(f'Error processing .xls file: {str(e)}', 'error')
                                return redirect(url_for('index'))

                    elif file.filename.endswith('.txt'):  # Handle .txt files (e.g., Unicode text)
                        try:
                            # Attempt to read the .txt file with appropriate encoding and delimiter
                            df_temp = pd.read_csv(file, delimiter='\t', encoding='utf-16')  # Adjust encoding/delimiter if needed
                            file_name = file.filename.rsplit('.', 1)[0]
                            csv_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_name}.csv")
                            df_temp.to_csv(csv_file_path, index=False, encoding='utf-8')  # Save as CSV
                            flash(f'Unicode text file converted to CSV and saved as {file_name}.csv', 'success')
                        except Exception as e:
                            flash(f'Error reading the .txt file: {str(e)}', 'error')
                            return redirect(url_for('index'))
                    ''''
                    first_column_name = df_temp.columns[0]
                    
                    if first_column_name != 'Date':
                        flash(f'Incorrect file uploaded! First column name should be "Date"', 'error')
                        return redirect(url_for('index'))
                    '''
                    # Capture file name for incorrect folder
                    file_name = file.filename.rsplit('.', 1)[0]

                    # adding this column
                    df_temp['Assignment'] = 'Orientation activity'

                    # Check if the file contains all required columns (including User)
                    if not all(column in df_temp.columns for column in required_columns):
                        missing_columns = [col for col in required_columns if col not in df_temp.columns]
                        flash(f'Missing required columns: {", ".join(missing_columns)}', 'error')
                        
                         
                         # Save the incorrect file (missing required column names) to the census_incorrect folder
                        incorrect_file_path = os.path.join(app.config['INCORRECT_FOLDER'], secure_filename(file.filename))
                        file.save(incorrect_file_path)
                        
                        # Save the file based on its type
                        if file.filename.endswith('.csv'):
                            df_temp.to_csv(incorrect_file_path, index=False)
                        elif file.filename.endswith('.xlsx'):
                            df_temp.to_excel(incorrect_file_path, index=False, sheet_name='Incorrect Data')

                        # Add a note to the file explaining why it was saved in the incorrect folder
                        try:
                            if file.filename.endswith('.xlsx'):
                                # Add a note to the Excel file
                                wb = load_workbook(incorrect_file_path)
                                ws = wb.active
                                last_row = ws.max_row
                                ws.cell(row=last_row + 2, column=1, value=f"Note: Missing required columns: {', '.join(missing_columns)}")
                                wb.save(incorrect_file_path)
                            elif file.filename.endswith('.csv'):
                                # Append a note to the CSV file
                                with open(incorrect_file_path, 'a') as f:
                                    f.write(f"\nNote: Missing required columns: {', '.join(missing_columns)}\n")
                        except Exception as e:
                            flash(f"Error adding note to incorrect file: {str(e)}", 'error')

                        flash(f'File saved in census_incorrect folder due to missing columns.', 'success')

                        return redirect(url_for('index'))
                     
                    
                    # Remove rows where the 'Name' column contains 'Test Student'
                    df_temp = df_temp[~df_temp['Name'].str.contains('Test Student', case=False, na=False)]
                
                    df = df_temp
                    flash('File upload successful!', 'success')
                    valid_file_uploaded = True  # Set the flag to True
                    
                    # Save the uploaded file with a unique name
                    original_file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
                    file.save(original_file_path)

                    return render_template('index.html', condition=True, valid_file_uploaded=valid_file_uploaded)
                
                except Exception as e:
                    flash(f'Error! File Type Incorrect! Save the file as a CSV file please.', 'error')
                    return redirect(url_for('index'))
            else:
                flash('Invalid file type! Please upload a csv, xls, xlsx, or txt file.', 'error')
        else:
            flash('No selected file', 'error')
    
    return render_template('index.html', condition=True, valid_file_uploaded=valid_file_uploaded)
'''
def save_incorrect_file(file, df_temp):
    incorrect_file_path = os.path.join(app.config['INCORRECT_FOLDER'], secure_filename(file.filename))
    df_temp.to_excel(incorrect_file_path, index=False, sheet_name='Census')
'''
# Route to delete columns
@app.route('/delete_columns', methods=['GET', 'POST'])
def delete_columns_page():
    global df
    if request.method == 'POST':
        if 'get_results' in request.form:
            # Get results button clicked, redirect to results page
            return redirect(url_for('results'))
        else:
            # Columns to keep
            kept_columns = COLUMNS_TO_KEEP
            # Columns to delete
            deleted_columns = [col for col in df.columns if col not in kept_columns]
            

            
            # Filter the dataframe to keep only the specified columns
            df = df[kept_columns]

            flash(f"Columns deleted successfully.", 'success')

            return render_template('delete_columns.html', deleted_columns=deleted_columns, kept_columns=kept_columns)

    return render_template('delete_columns.html', deleted_columns=[], kept_columns=[])

# Delete rows route
@app.route('/delete_rows', methods=['GET', 'POST'])
def delete_rows_page():
    global df
    if request.method == 'POST':
        column_name = request.form['column_name']
        value_to_delete = request.form['value_to_delete']
        df = delete_rows_by_value(df, column_name, value_to_delete)
        flash(f"Rows with '{value_to_delete}' in '{column_name}' deleted.", 'success')
        return redirect(url_for('results', save=save))

    # Pass kept columns for row deletion options
    kept_columns = request.args.getlist('kept_columns')
    return render_template('delete_rows.html', columns=kept_columns)


# results page
@app.route('/results', methods=['GET', 'POST'])
def results():
    global df

    # Changing 2 column names per request*
    df = df.rename(columns={'Submitted': 'Date Submitted', 'OverallScore': 'Score'})

    # Remove rows where the 'Name' column contains the word 'Test Student' (case-insensitive)
    df_cleaned = df[~df['Name'].str.contains('Test Student', case=False, na=False)].dropna()

    # Sort the dataframe by 'id' and 'score' columns
    # 'score' in descending order ensures the highest grade comes first
    df_cleaned = df_cleaned.sort_values(by=['SISID','Score'], ascending=[True, False])# sort by value
    # Drop duplicates, keeping the first occurrence (which now has the highest 'Score')
    df_cleaned = df_cleaned.drop_duplicates(subset='SISID', keep='first')

    # Reset the index after cleaning and dropping duplicates (optional)
    df_cleaned.reset_index(drop=True, inplace=True)
        
    # Sort the DataFrame by the 'name' column
    df_sorted = df_cleaned.sort_values(by='Name')

    # Count the number of unique users
    user_count = df_sorted['SISID'].nunique()

    if request.method == 'POST':
        filename = request.form['filename']
        if filename:
            return redirect(url_for('download', filename=filename))

    return render_template('results.html', df_updated=df_sorted.to_html(), user_count=user_count,save = save)


# download file as excel
@app.route('/download', methods=['POST'])
def download():
    global df
    #filename = request.form['filename']
    filename = file_name

    # Limit "Column" values to 20 characters
    #df['Column'] = df['Column'].apply(lambda x: x[:20] if isinstance(x, str) else x)

    census_assign1 = 'Orientation activity'
    cencus_assign2 = 'Census Entry Quiz'

    # Find the first index where the first column has a value
    #first_valid_index = df['Date'].first_valid_index()

    # Convert the 'Value' column to numeric, coercing errors to NaN
    df['Score'] = pd.to_numeric(df['Score'], errors='coerce')

    # Drop rows where 'Value' is NaN
    df_cleaned = df.dropna(subset=['Score'])

    # Remove rows where the 'name' column contains 'Preview' or 'student' (case-insensitive)
    df_cleaned = df_cleaned[~df_cleaned['Name'].str.contains('Test Student', case=False, na=False)]

    # Sort by 'Username' and 'Value' (descending for Value to keep highest grades)
    df_cleaned = df_cleaned.sort_values(by=['SISID', 'Score'], ascending=[True, False])

    # Drop duplicates, keeping the one with the highest 'Value'
    df_cleaned = df_cleaned.drop_duplicates(subset='SISID', keep='first')

    # in canvas there is no column named column ************************

   # Check if the column contains only one of the two specified values
    mask = df_cleaned['Assignment'].isin([census_assign1])

    # Check if all values under the column are the same
    all_same = df_cleaned['Assignment'].eq(df_cleaned['Assignment'].iloc[0]).all()

    # Combine the two conditions
    result = mask.all() or all_same

    # Filter rows where the grade condition is met
    # Check for a specific value
    grade = df_cleaned['Score'] >=1

    # now check
    if not result: # not true
        
        # show results
        result="Orientation activity Invalid or NOT consistent  !  "
        df_sorted, user_count, save, file = census_incorrect(filename, result)
    
    elif not grade.all(): # grade condition not tru

        # show results
        result="One or multiple students do not have a grade!"
        df_sorted, user_count, save, file = census_incorrect(filename, result)
    
    else:
        
        # Sort the DataFrame by the 'User' column
        df_sorted = df_cleaned.sort_values(by='Name')
        # Reset the index (optional)
        df_sorted.reset_index(drop=True, inplace=True)

        # Count the number of unique username
        user_count = df_sorted['SISID'].nunique()
        count_str = str(user_count)
        
        # Combine the current directory with the new directory name
        download_path = os.path.join(current_directory, DOWNLOAD_FOLDER)


        # Create the new directory
        os.makedirs(download_path, exist_ok=True)

        # Ensure the downloads folder exists
        app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER

        # Write the count to column D (4th column, hence D is column 3 in zero-indexed system)
        #worksheet.write(count_row, 3, 'TOTAL STUDENT COUNT: ' + count_str)

        download_name=f"{filename}.xlsx"

        file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], download_name)

        # Save the file to the specified folder
        df_sorted.to_excel(file_path, sheet_name='Census', index=False, header=True)

        # Reopen the file to write additional information
        wb = load_workbook(file_path)
        ws = wb['Census']

        # Set page layout to landscape
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Write additional information at the bottom
        last_row = ws.max_row
        info_row = last_row + 3
        ws.cell(row=info_row, column=3, value='TOTAL STUDENT COUNT: ' + str(df_sorted['SISID'].nunique()))

        # Set the header for printing
        ws.oddHeader.center.text = filename

        # Save the workbook
        wb.save(file_path)

        file = f'{file_path}'
        save = f'File successfully created. See Path.. '

    return render_template('results.html', df_updated=df_sorted.to_html(), user_count=user_count, save=save, file=file)
    
def census_incorrect(filename, result):
    global df
    
    # Limit "Column" values to 20 characters
    #df['Column'] = df['Column'].apply(lambda x: x[:20] if isinstance(x, str) else x)

   # Find the first index where the first column has a value
    #first_valid_index = df['Date'].first_valid_index()

    # Convert the 'Value' column to numeric, coercing errors to NaN
    df['Score'] = pd.to_numeric(df['Score'], errors='coerce')

    # Drop rows where 'Value' is NaN
    df_cleaned = df.dropna(subset=['Score'])

    # Remove rows where the 'User' column contains 'Preview' or 'student' (case-insensitive)
    df_cleaned = df_cleaned[~df_cleaned['Name'].str.contains('Preview|student', case=False, na=False)]

    # Sort by 'Username' and 'Value' (descending for Value to keep highest grades)
    df_cleaned = df_cleaned.sort_values(by=['SISID', 'Score'], ascending=[True, False])

    # Drop duplicates, keeping the one with the highest 'Value'
    df_cleaned = df_cleaned.drop_duplicates(subset='SISID', keep='first')


    # Reset the index (optional)
    df_cleaned.reset_index(drop=True, inplace=True)

    # Sort the DataFrame by the 'User' column
    df_sorted = df_cleaned.sort_values(by='Name') 

    # Count the number of unique users
    user_count = df_sorted['SISID'].nunique()
    count_str = str(user_count)
        
    # Combine the current directory with the new directory name
    download_path = os.path.join(current_directory, INCORRECT_FOLDER)


    # Create the new directory
    os.makedirs(download_path, exist_ok=True)

    # Ensure the downloads folder exists
    app.config['INCORRECT_FOLDER'] = INCORRECT_FOLDER

    download_name=f"{filename}.xlsx"

    file_path = os.path.join(app.config['INCORRECT_FOLDER'], download_name)

    # Save the file to the specified folder
    df_sorted.to_excel(file_path, sheet_name='Census', index=False)

    # Reopen the file to write additional information
    wb = load_workbook(file_path)
    ws = wb['Census']

    # Set page layout to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Find the last row with data
    last_row = ws.max_row

    # Write information 3 rows below the last row
    info_row = last_row + 3
    ws.cell(row=info_row, column=3, value='TOTAL STUDENT COUNT: ' + count_str)

    # Add the note explaining why the file is incorrect
    note_row = info_row + 2
    ws.cell(row=note_row, column=1, value="Note:")
    ws.cell(row=note_row + 1, column=1, value=result)

    # Set the header for printing
    ws.oddHeader.center.text = file_name
    
    # Save the workbook
    wb.save(file_path)
        
    # show results
    file = f'File saved. See Path.. '+f' {file_path}'

    save = result
    #save = f'Orientation activity Invalid or NOT consistent  !  '
    
    return df_sorted, user_count, save, file
    
# Helper functions
def delete_rows_by_value(df, column_name, value_to_delete):
    if column_name in df.columns:
        condition = df[column_name] == value_to_delete
        df = df.drop(df[condition].index)
    return df

if __name__ == '__main__':
    app.run(debug=True)

    
    #webview.start()
